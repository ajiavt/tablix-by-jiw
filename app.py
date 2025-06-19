import os
import re
import json
import pandas as pd
import threading
import concurrent.futures
import time
import math
from queue import Queue
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, abort
from werkzeug.utils import secure_filename
from flask_cors import CORS  # Tambahkan import untuk CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import os.path
import shutil
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

app = Flask(__name__) 
CORS(app)  # Aktifkan CORS untuk semua route

app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['SOURCE_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'source')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload size
app.config['MAX_WORKERS'] = 4  # Jumlah thread untuk pencarian paralel
app.config['SEARCH_ACTIVE'] = False  # Flag untuk status pencarian aktif
app.config['ALLOWED_EXTENSIONS'] = {'dat', 'DAT', 'txt', 'TXT', 'sql', 'SQL'}

# Cek folder source secara default
default_source_path = os.path.join(os.getcwd(), 'source')
if os.path.exists(default_source_path) and os.path.isdir(default_source_path):
    app.config['SOURCE_FOLDER'] = default_source_path
    print(f"Using default source folder: {default_source_path}")

# Cek folder intellij script secara default
DEFAULT_INTELLIJ_PATH = "/Users/ajiavt/IdeaProjects/olibs724-be-kv-sumut/src/main/webapp/script/"
if os.path.exists(DEFAULT_INTELLIJ_PATH) and os.path.isdir(DEFAULT_INTELLIJ_PATH):
    app.config['INTELLIJ_PATH'] = DEFAULT_INTELLIJ_PATH
    print(f"Using default Intellij script path: {DEFAULT_INTELLIJ_PATH}")

# Inisialisasi variabel
SEARCH_RESULTS = []
SEARCH_RESULTS_LOCK = threading.Lock()
CANCEL_SEARCH = False

# Variabel untuk menyimpan hasil ekspor Excel
EXCEL_RESULTS = []
EXCEL_RESULTS_LOCK = threading.Lock()

# Variabel untuk lacak status pencarian
SEARCH_ACTIVE = False
SEARCH_COMPLETE = False
SEARCH_TOTAL_FILES = 0
SEARCH_PROCESSED_FILES = 0

# Baca file tbl_name.txt untuk mendapatkan daftar tabel
def get_table_names():
    """Membaca file setting-tbl-name.txt dan mengembalikan daftar nama tabel dengan remote server-nya."""
    try:
        table_names = []
        
        with open('setting-tbl-name.txt', 'r') as file:
            for line in file:
                line = line.strip()
                if not line:
                    continue
                
                parts = line.split('\t')  # File menggunakan tab sebagai separator
                if len(parts) < 2:
                    continue
                    
                table_name = parts[0].strip()
                remote_server = parts[1].strip() if len(parts) > 1 else ""
                new_table_name = parts[2].strip() if len(parts) > 2 else table_name
                
                # Tambahkan ke list hasil
                table_names.append({
                    "table_name": table_name,
                    "remote_server": remote_server,
                    "new_table_name": new_table_name
                })
                
        return table_names
    except Exception as e:
        print(f"Error saat membaca setting-tbl-name.txt: {e}")
        return []

# Baca file script-name.txt untuk mendapatkan nama script
def get_script_name(file_name):
    """Mendapatkan nama script dari file name."""
    script_names = get_script_names()
    return script_names.get(file_name, "")

def get_script_names():
    try:
        script_names = {}
        with open('setting-script-name.txt', 'r') as file:
            for line in file:
                line = line.strip()
                if not line:
                    continue
                
                # Coba split dengan tab dulu
                parts = line.split('\t')
                
                if len(parts) == 1:
                    # Jika tidak ada tab, coba split dengan spasi
                    parts = line.split(None, 1)  # Split pada whitespace pertama untuk memisahkan kode dan nama
                
                if len(parts) >= 2:
                    script_code = parts[0].strip()
                    script_name = parts[1].strip()
                    script_names[script_code] = script_name
        
        return script_names
    except Exception as e:
        print(f"Error membaca setting-script-name.txt: {e}")
        return {}

def load_tbl_be_tables():
    """Load tabel BE dari file setting-tbl-be.txt"""
    try:
        with open('setting-tbl-be.txt', 'r') as file:
            tables = [line.strip() for line in file.readlines() if line.strip()]
        return tables
    except Exception as e:
        print(f"Error loading BE tables: {e}")
        return []

# Variabel global untuk data tabel
TABLE_INFO = get_table_names()

# Map tabel ke remote server
table_to_server = {table_info['table_name']: table_info['remote_server'] for table_info in TABLE_INFO}

# Load daftar tabel BE untuk pengecekan difficulty
all_be_tables = load_tbl_be_tables()

# Pastikan folder uploads ada
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
# Pastikan folder source ada
os.makedirs(app.config['SOURCE_FOLDER'], exist_ok=True)

# Keyword default yang disediakan
DEFAULT_KEYWORDS = ["jt", "jdbc", "jdbcTemplate"]

# Pre-compile regex pattern yang sering digunakan
METHOD_PATTERN = re.compile(r'(?:public|private|protected)?\s+(?:\w+\s+)(\w+)\s*\([^)]*\)\s*\{([^{}]*(?:\{[^{}]*\}[^{}]*)*)\}', re.DOTALL)

# Fungsi untuk menganalisis file tertentu dengan fokus pada statement tunggal
def analyze_file(file_path, table_name, keywords=None):
    global CANCEL_SEARCH, SEARCH_RESULTS, SEARCH_RESULTS_LOCK
    
    # Cek apakah pencarian dibatalkan
    if CANCEL_SEARCH:
        return []
    
    results = []
    found_statements = set()  # Untuk mencegah duplikasi
    
    # Baca informasi remote server dari setting-remote-server.txt
    remote_server_map = {}
    try:
        with open('setting-remote-server.txt', 'r') as rs_file:
            for line in rs_file:
                line = line.strip()
                if line:
                    parts = line.split(' ', 1)
                    if len(parts) > 1:
                        remote_server_map[parts[0]] = parts[1]
    except Exception as e:
        print(f"Error membaca setting-remote-server.txt: {e}")
    
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()
            
            # Jika nama tabel tidak ada dalam konten file, langsung return
            # Gunakan regex dengan word boundary untuk exact match
            # Pattern ini akan mencocokkan tabel jika:
            # 1. Di awal string atau diawali pembatas (spasi, tanda kutip, dsb)
            # 2. Di akhir string atau diakhiri pembatas (spasi, tanda kutip, dsb)
            table_pattern = r'(^|[\s\'"`\(\),;])' + re.escape(table_name) + r'($|[\s\'"`\(\),;])'
            if not re.search(table_pattern, content, re.IGNORECASE):
                return []
            
            # Jika ada keyword, periksa apakah juga ada dalam konten
            keyword_found_in_file = False
            if keywords:
                for keyword in keywords:
                    # Jika keyword adalah '-', itu berarti kita tidak memerlukan keyword khusus
                    if keyword == '-':
                        keyword_found_in_file = True
                        break
                    
                    if keyword.lower() in content.lower():
                        keyword_found_in_file = True
                        break
                
                if not keyword_found_in_file:
                    return []
            
            # Bagi file menjadi baris-baris
            lines = content.split('\n')
            method_name = os.path.basename(file_path).replace('.DAT', '')
            
            # Dapatkan remote server dari nama file
            remote_server = ""
            for rs_name in remote_server_map.keys():
                if rs_name in content:
                    remote_server = rs_name
                    break
            
            # Multi-baris statement tracking
            i = 0
            while i < len(lines):
                # Cek pembatalan
                if CANCEL_SEARCH:
                    return []
                
                line = lines[i].strip()
                if not line:  # Skip baris kosong
                    i += 1
                    continue
                
                line_lower = line.lower()
                
                # 1. Deteksi SQL query dengan tabel di dalamnya
                sql_indicators = [
                    'select', 'from', 'update', 'insert into', 'delete from', 
                    'jt.query', 'jt.update', 'jdbctemplate', 'createquery'
                ]
                
                is_potential_sql = any(ind in line_lower for ind in sql_indicators)
                
                # Cek apakah baris saat ini berisi nama tabel dengan exact match
                table_in_line = False
                table_pattern_line = r'(^|[\s\'"`\(\),;])' + re.escape(table_name) + r'($|[\s\'"`\(\),;])'
                if re.search(table_pattern_line, line, re.IGNORECASE):
                    table_in_line = True
                
                if table_in_line or (is_potential_sql and i + 10 < len(lines)):
                    # Kumpulkan beberapa baris untuk menangkap statement lengkap
                    # Kasus SQL multi-baris dengan concat string
                    statement_lines = [line]
                    j = i + 1
                    quote_open = line.count('"') % 2 == 1 or line.count("'") % 2 == 1
                    brackets_count = line.count('(') - line.count(')')
                    continued_line = line.endswith('+') or line.strip().endswith("\\")
                    
                    # Batas jumlah baris untuk menghindari pengecekan terlalu banyak baris
                    max_lines = 20
                    line_count = 1
                    
                    # Lacak statement multi-baris
                    while (j < len(lines) and line_count < max_lines and 
                           (quote_open or brackets_count > 0 or continued_line)):
                        next_line = lines[j].strip()
                        if not next_line:  # Skip baris kosong
                            j += 1
                            continue
                        
                        statement_lines.append(next_line)
                        line_count += 1
                        
                        # Update state pelacakan
                        for char in next_line:
                            if char == '"' and not next_line.endswith('\\"'):
                                quote_open = not quote_open
                            if char == "'":
                                quote_open = not quote_open
                        
                        brackets_count += next_line.count('(') - next_line.count(')')
                        continued_line = next_line.strip().endswith('+') or next_line.strip().endswith("\\")
                        
                        # Jika statement sudah lengkap (tidak ada tanda lanjutan, kurung sudah lengkap, dan quote sudah ditutup)
                        if not quote_open and brackets_count <= 0 and not continued_line:
                            break
                        
                        j += 1
                    
                    # Gabungkan statement lengkap
                    statement_text = '\n'.join(statement_lines)
                    statement_text_lower = statement_text.lower()
                    
                    # 2. Sekarang periksa apakah statement mengandung tabel yang dicari dengan exact match
                    # Untuk SQL query, tabel bisa muncul setelah FROM, JOIN, UPDATE, INTO, dll.
                    table_in_statement = False
                    table_pattern_stmt = r'(^|[\s\'"`\(\),;])' + re.escape(table_name) + r'($|[\s\'"`\(\),;])'
                    if re.search(table_pattern_stmt, statement_text, re.IGNORECASE):
                        table_in_statement = True
                    
                    if table_in_statement:
                        matched_keyword = ""
                        keyword_matched = False
                        
                        # Periksa apakah statement mengandung keyword
                        if keywords:
                            for keyword in keywords:
                                # Jika keyword adalah '-', itu berarti kita menerima tanpa keyword
                                if keyword == '-':
                                    keyword_matched = True
                                    matched_keyword = "-"
                                    break
                                
                                if keyword.lower() in statement_text_lower:
                                    keyword_matched = True
                                    matched_keyword = keyword
                                    break
                        else:
                            # Jika tidak ada keyword yang ditentukan, setiap statement dengan tabel dianggap cocok
                            keyword_matched = True
                        
                        # Tambahkan ke hasil jika statement valid dan sesuai dengan kriteria
                        if keyword_matched:
                            # Tambahkan ke hasil jika belum ada statement yang sama
                            statement_hash = hash(statement_text)
                            if statement_hash not in found_statements:
                                found_statements.add(statement_hash)
                                
                                # Periksa apakah perlu direplace
                                needs_replace = False
                                if matched_keyword:
                                    keyword_lower = matched_keyword.lower()
                                    if (keyword_lower == 'jt' or 
                                        keyword_lower == 'jdbctemplate' or 
                                        'getjdbctemplate' in keyword_lower or 
                                        'jdbc' in keyword_lower):
                                        
                                        # Hanya tandai butuh replace jika statement mengandung pattern replace yang valid
                                        statement_lower = statement_text.lower()
                                        if 'jt.' in statement_lower or 'jdbctemplate' in statement_lower or 'getjdbctemplate' in statement_lower:
                                            needs_replace = True
                                
                                # Tentukan jenis query (SELECT atau manipulasi data)
                                query_type = []
                                statement_lower = statement_text.lower()
                                
                                # Cek untuk SELECT
                                # 1. Logika utama - mengandung kata 'select'
                                if 'select' in statement_lower:
                                    query_type.append('select')
                                
                                # 2. Logika pengecekan method query
                                query_methods = [
                                    '.query(', 
                                    'queryobject(', 
                                    'queryforobject(',
                                    'createquery(',
                                    'queryforint(',
                                    'queryforstring(',
                                    'queryforbigdecimal(',
                                    '.queryforobject(',
                                    '.queryforint(',
                                    '.queryforstring(',
                                    '.queryforbigdecimal(',
                                ]
                                
                                for method in query_methods:
                                    if method in statement_lower:
                                        query_type.append('select')
                                        break  # Hentikan loop jika sudah ditemukan
                                
                                # 3. Tambahan untuk metode dengan string SELECT langsung
                                if ('jt.query("select' in statement_lower or 
                                    "jt.query('select" in statement_lower or
                                    'queryforobject(("select' in statement_lower or 
                                    "queryforobject(('select" in statement_lower or
                                    'queryforobject("select' in statement_lower or 
                                    "queryforobject('select" in statement_lower):
                                    query_type.append('select')
                                
                                # Cek untuk manipulasi data (DML)
                                if (re.search(r'\binsert\s+into\b', statement_lower) or 
                                    re.search(r'\bupdate\b', statement_lower) or 
                                    re.search(r'\bdelete\s+from\b', statement_lower) or
                                    'insertdataquery' in statement_lower or
                                    'updatedataquery' in statement_lower or
                                    'deletedataquery' in statement_lower or
                                    'insertdata' in statement_lower or
                                    'updatedata' in statement_lower or
                                    'deletedata' in statement_lower or
                                    'jt.update(' in statement_lower or
                                    'jt.insert(' in statement_lower or
                                    'jt.delete(' in statement_lower or
                                    'jt.query(jt.insertdataquery' in statement_lower or
                                    'jt.addsql(jt.insertdataquery' in statement_lower or
                                    'jt.query(jt.updatedataquery' in statement_lower or
                                    'jt.addsql(jt.updatedataquery' in statement_lower or
                                    'jt.query(jt.deletedataquery' in statement_lower or
                                    'jt.addsql(jt.deletedataquery' in statement_lower or
                                    'jt.addsql("update' in statement_lower or
                                    'jt.addsql("delete' in statement_lower or
                                    'jt.addsql("insert' in statement_lower or
                                    "jt.addsql('update" in statement_lower or
                                    "jt.addsql('delete" in statement_lower or
                                    "jt.addsql('insert" in statement_lower or
                                    'jt.execute("update' in statement_lower or
                                    'jt.execute("delete' in statement_lower or
                                    'jt.execute("insert' in statement_lower or
                                    "jt.execute('update" in statement_lower or
                                    "jt.execute('delete" in statement_lower or
                                    "jt.execute('insert" in statement_lower):
                                    query_type.append('manipulation')
                                
                                # Jika tidak ada tipe yang terdeteksi, kategorikan sebagai 'other'
                                if not query_type:
                                    query_type = ['other']
                                
                                result = {
                                    "file_name": os.path.basename(file_path),
                                    "method_name": method_name,
                                    "table_name": table_name,
                                    "remote_server": remote_server,
                                    "query": statement_text,
                                    "keyword": matched_keyword,
                                    "line_number": i + 1,  # +1 karena indeks baris dimulai dari 0
                                    "needs_replace": needs_replace,
                                    "replaced": False,  # Awalnya belum direplace
                                    "query_type": query_type  # Jenis query sebagai array
                                }
                                
                                results.append(result)
                                
                                # Tambahkan ke hasil dengan lock
                                with SEARCH_RESULTS_LOCK:
                                    SEARCH_RESULTS.append(result)
                
                # 1. Cek pattern .put() method yang umum digunakan di Java map
                elif ".put(" in line_lower:
                    # Ini adalah kasus Map.put() yang mungkin mengandung keyword
                    # Cek apakah tabel terdapat dalam baris dengan exact match
                    table_pattern_put = r'(^|[\s\'"`\(\),;])' + re.escape(table_name) + r'($|[\s\'"`\(\),;])'
                    if re.search(table_pattern_put, line, re.IGNORECASE):
                        statement_lines = [line]
                        
                        # Cek apakah ini mengandung keyword
                        keyword_found = False
                        matched_keyword = ""
                        
                        if keywords:
                            for keyword in keywords:
                                # Jika keyword adalah '-', itu berarti kita menerima tanpa keyword
                                if keyword == '-':
                                    keyword_found = True
                                    matched_keyword = "-"
                                    break
                                
                                if keyword.lower() in line_lower:
                                    keyword_found = True
                                    matched_keyword = keyword
                                    break
                        else:
                            keyword_found = True  # Anggap ditemukan jika tidak ada keyword yang dicari
                        
                        if keyword_found:
                            statement_text = '\n'.join(statement_lines)
                            statement_hash = hash(statement_text)
                            if statement_hash not in found_statements:
                                found_statements.add(statement_hash)
                                
                                # Periksa apakah perlu direplace
                                needs_replace = False
                                if matched_keyword:
                                    keyword_lower = matched_keyword.lower()
                                    if (keyword_lower == 'jt' or 
                                        keyword_lower == 'jdbctemplate' or 
                                        'getjdbctemplate' in keyword_lower or 
                                        'jdbc' in keyword_lower):
                                        
                                        # Hanya tandai butuh replace jika statement mengandung pattern replace yang valid
                                        statement_lower = statement_text.lower()
                                        if 'jt.' in statement_lower or 'jdbctemplate' in statement_lower or 'getjdbctemplate' in statement_lower:
                                            needs_replace = True
                                
                                # Method put tidak termasuk SELECT atau manipulasi data
                                query_type = ['other']
                                
                                result = {
                                    "file_name": os.path.basename(file_path),
                                    "method_name": method_name,
                                    "table_name": table_name,
                                    "remote_server": remote_server,
                                    "query": statement_text,
                                    "keyword": matched_keyword,
                                    "line_number": i + 1,  # +1 karena indeks baris dimulai dari 0
                                    "needs_replace": needs_replace,
                                    "replaced": False,  # Awalnya belum direplace
                                    "query_type": query_type
                                }
                                
                                results.append(result)
                                
                                # Tambahkan ke hasil dengan lock
                                with SEARCH_RESULTS_LOCK:
                                    SEARCH_RESULTS.append(result)
                
                # 2. Deteksi statement Java yang berisi nama tabel dan keyword                
                elif True:
                    # Cek apakah tabel terdapat dalam baris dengan exact match
                    table_pattern_java = r'(^|[\s\'"`\(\),;])' + re.escape(table_name) + r'($|[\s\'"`\(\),;])'
                    if not re.search(table_pattern_java, line, re.IGNORECASE):
                        i += 1
                        continue
                        
                    # Jika ada keyword, periksa apakah dalam statement yang sama
                    if keywords:
                        keyword_found = False
                        matched_keyword = ""
                        
                        for keyword in keywords:
                            # Jika keyword adalah '-', itu berarti kita menerima tanpa keyword
                            if keyword == '-':
                                keyword_found = True
                                matched_keyword = "-"
                                break
                            
                            if keyword.lower() in line_lower:
                                keyword_found = True
                                matched_keyword = keyword
                                break
                        
                        if not keyword_found:
                            # Mencari keyword dalam 3 baris ke depan atau ke belakang
                            # hanya jika ada indikasi metode JDBC
                            jdbc_indicators = [
                                'query(', '.execute(', '.update(', '.insert(', '.delete(', 'addsql(', 
                                'insertdataquery(', 'updatedataquery(', 'queryobject(', 'deletedata('
                            ]
                            
                            is_jdbc_call = any(ind in line_lower for ind in jdbc_indicators)
                            
                            if is_jdbc_call:
                                # Cari keyword dalam beberapa baris di sekitarnya
                                start_look = max(0, i - 3)  # Check farther back for context
                                end_look = min(len(lines), i + 4)
                                
                                for j in range(start_look, end_look):
                                    if j == i:  # Skip baris saat ini
                                        continue
                                    
                                    check_line = lines[j].strip().lower()
                                    for keyword in keywords:
                                        # Jika keyword adalah '-', itu berarti kita menerima tanpa keyword
                                        if keyword == '-':
                                            keyword_found = True
                                            matched_keyword = "-"
                                            break
                                        
                                        if keyword.lower() in check_line:
                                            # Periksa apakah ini adalah statement yang berhubungan
                                            if ('=' in line_lower or ';' in line_lower or 
                                                '(' in line_lower or ')' in line_lower or 
                                                '.put(' in check_line):  # Tambahan untuk deteksi .put()
                                                keyword_found = True
                                                matched_keyword = keyword
                                                
                                                # Jika keyword ditemukan di baris dengan method .put()
                                                # prioritaskan baris tersebut
                                                if '.put(' in check_line and table_name.lower() in check_line:
                                                    i = j  # Pindah ke baris ini
                                                    line = lines[j].strip()
                                                    line_lower = line.lower()
                                                break
                                    
                                    if keyword_found:
                                        break
                            
                        if not keyword_found:
                            i += 1
                            continue
                    else:
                        matched_keyword = ""
                    
                    # Ini adalah statement yang valid - dengan tabel dan keyword (jika ada)
                    # Cari batas dari statement Java ini (biasanya diakhiri dengan ;)
                    statement_lines = []
                    
                    # Ini adalah kasus khusus untuk metode JDBC dengan tabel di dalamnya
                    jdbc_call_indicators = [
                        '.query(', '.execute(', '.update(', '.insert(', '.delete(', '.addsql(', 
                        'insertdataquery(', 'updatedataquery(', 'queryobject(', '.deletedata(',
                        'jt.', '.put('
                    ]
                    
                    is_jdbc_call = any(ind in line_lower for ind in jdbc_call_indicators)
                    
                    if is_jdbc_call:
                        statement_lines.append(line)
                        
                        # Beberapa pernyataan JDBC mungkin bisa dalam satu baris
                        if ';' in line:
                            pass  # Statement selesai dalam satu baris
                        else:
                            # Cari baris berikutnya sampai statement berakhir
                            j = i + 1
                            open_brackets = line.count('(') - line.count(')')
                            
                            while j < len(lines) and open_brackets > 0:
                                next_line = lines[j].strip()
                                if next_line:  # Skip baris kosong
                                    statement_lines.append(next_line)
                                    open_brackets += next_line.count('(') - next_line.count(')')
                                    
                                    if ';' in next_line and open_brackets <= 0:
                                        break
                                
                                j += 1
                    else:
                        # Bukan JDBC call, mungkin hanya assignment atau deklarasi
                        # Ambil hanya baris yang berisi tabel dan keyword saja
                        statement_lines.append(line)
                    
                    # Gabungkan ke statement lengkap
                    statement_text = '\n'.join(statement_lines)
                    
                    # Tambahkan ke hasil jika belum ada statement yang sama
                    statement_hash = hash(statement_text)
                    if statement_hash not in found_statements:
                        found_statements.add(statement_hash)
                        
                        # Filter statement yang tidak relevan atau noise
                        if len(statement_lines) > 0 and table_name.lower() in statement_text.lower():
                            # Terima statement yang berkaitan dengan operasi pada objek tabel
                            is_valid_statement = (
                                ('jt.' in statement_text.lower() and '(' in statement_text and ')' in statement_text) or
                                ('.put(' in statement_text.lower() and table_name.lower() in statement_text.lower()) or 
                                ('insert' in statement_text.lower() and table_name.lower() in statement_text.lower()) or
                                ('update' in statement_text.lower() and table_name.lower() in statement_text.lower()) or
                                ('delete' in statement_text.lower() and table_name.lower() in statement_text.lower()) or
                                ('select' in statement_text.lower() and table_name.lower() in statement_text.lower())
                            )
                            
                            if is_valid_statement:
                                # Periksa apakah perlu direplace
                                needs_replace = False
                                if matched_keyword:
                                    keyword_lower = matched_keyword.lower()
                                    if (keyword_lower == 'jt' or 
                                        keyword_lower == 'jdbctemplate' or 
                                        'getjdbctemplate' in keyword_lower or 
                                        'jdbc' in keyword_lower):
                                        
                                        # Hanya tandai butuh replace jika statement mengandung pattern replace yang valid
                                        statement_lower = statement_text.lower()
                                        if 'jt.' in statement_lower or 'jdbctemplate' in statement_lower or 'getjdbctemplate' in statement_lower:
                                            needs_replace = True
                                
                                # Tentukan jenis query (SELECT atau manipulasi data)
                                query_type = []
                                statement_lower = statement_text.lower()
                                
                                # Cek untuk SELECT
                                # 1. Logika utama - mengandung kata 'select'
                                if 'select' in statement_lower:
                                    query_type.append('select')
                                
                                # 2. Logika pengecekan method query
                                query_methods = [
                                    '.query(', 
                                    'queryobject(', 
                                    'queryforobject(',
                                    'createquery(',
                                    'queryforint(',
                                    'queryforstring(',
                                    'queryforbigdecimal(',
                                    '.queryforobject(',
                                    '.queryforint(',
                                    '.queryforstring(',
                                    '.queryforbigdecimal(',
                                ]
                                
                                for method in query_methods:
                                    if method in statement_lower:
                                        query_type.append('select')
                                        break  # Hentikan loop jika sudah ditemukan
                                
                                # 3. Tambahan untuk metode dengan string SELECT langsung
                                if ('jt.query("select' in statement_lower or 
                                    "jt.query('select" in statement_lower or
                                    'queryforobject(("select' in statement_lower or 
                                    "queryforobject(('select" in statement_lower or
                                    'queryforobject("select' in statement_lower or 
                                    "queryforobject('select" in statement_lower):
                                    query_type.append('select')
                                
                                # Cek untuk manipulasi data
                                if ('.update(' in statement_lower or 
                                    '.insert(' in statement_lower or 
                                    '.delete(' in statement_lower or 
                                    'update' in statement_lower or 
                                    'insert into' in statement_lower or 
                                    'delete from' in statement_lower or
                                    'insertdataquery' in statement_lower or
                                    'updatedataquery' in statement_lower or
                                    'deletedataquery' in statement_lower or
                                    'insertdata' in statement_lower or
                                    'updatedata' in statement_lower or
                                    'deletedata' in statement_lower or
                                    'jt.query(jt.insertdataquery' in statement_lower or
                                    'jt.addsql(jt.insertdataquery' in statement_lower or
                                    'jt.query(jt.updatedataquery' in statement_lower or
                                    'jt.addsql(jt.updatedataquery' in statement_lower or
                                    'jt.query(jt.deletedataquery' in statement_lower or
                                    'jt.addsql(jt.deletedataquery' in statement_lower or
                                    'jt.addsql("update' in statement_lower or
                                    'jt.addsql("delete' in statement_lower or
                                    'jt.addsql("insert' in statement_lower or
                                    "jt.addsql('update" in statement_lower or
                                    "jt.addsql('delete" in statement_lower or
                                    "jt.addsql('insert" in statement_lower or
                                    'jt.execute("update' in statement_lower or
                                    'jt.execute("delete' in statement_lower or
                                    'jt.execute("insert' in statement_lower or
                                    "jt.execute('update" in statement_lower or
                                    "jt.execute('delete" in statement_lower or
                                    "jt.execute('insert" in statement_lower):
                                    query_type.append('manipulation')
                                
                                # Jika tidak ada tipe yang terdeteksi, kategorikan sebagai 'other'
                                if not query_type:
                                    query_type = ['other']
                                
                                result = {
                                    "file_name": os.path.basename(file_path),
                                    "method_name": method_name,
                                    "table_name": table_name,
                                    "remote_server": remote_server,
                                    "query": statement_text,
                                    "keyword": matched_keyword,
                                    "line_number": i + 1,  # +1 karena indeks baris dimulai dari 0
                                    "needs_replace": needs_replace,
                                    "replaced": already_replaced,  # Set replaced = True jika sudah direplace
                                    "query_type": query_type,
                                    "difficulty": difficulty
                                }
                                
                                results.append(result)
                                
                                # Tambahkan ke hasil dengan lock
                                with SEARCH_RESULTS_LOCK:
                                    SEARCH_RESULTS.append(result)
                
                i += 1
    
    except Exception as e:
        print(f"Error pada file {file_path}: {e}")
    
    return results

# Fungsi untuk mencari query dalam file
def search_queries(table_name, keywords, source_folder, skip_js=False, num_partitions=1):
    global CANCEL_SEARCH, SEARCH_RESULTS, EXCEL_RESULTS
    
    if not table_name:
        return {"status": "error", "message": "Nama tabel harus diisi"}
    
    # Ambil semua file .DAT di folder source
    all_files = []
    for root, _, filenames in os.walk(source_folder):
        for filename in filenames:
            if filename.endswith('.DAT'):
                if skip_js and 'JS' in filename:
                    continue
                all_files.append(os.path.join(root, filename))
    
    if not all_files:
        return {"status": "error", "message": "Tidak ada file .DAT di folder source"}
    
    print(f"[{table_name}] Analyzing {len(all_files)} files")
    
    local_results = []
    
    for file_path in all_files:
        # Jika pencarian dibatalkan, hentikan
        if CANCEL_SEARCH:
            break
        
        # Lakukan analisis file, menambahkan hasilnya langsung ke SEARCH_RESULTS dan EXCEL_RESULTS
        # menggunakan thread locking di dalam fungsi analyze_file
        file_results = analyze_file(file_path, table_name, keywords)
        local_results.extend(file_results)
    
    print(f"[{table_name}] Search completed: {len(local_results)} results found")
    
    return {"status": "success", "count": len(local_results), "results": local_results}

# Route untuk halaman utama
@app.route('/')
def index():
    table_info = get_table_names()
    script_names = get_script_names()  # Tambahkan ini
    return render_template('index.html', table_info=table_info, default_keywords=DEFAULT_KEYWORDS, script_names=script_names)

# Route untuk mendapatkan jumlah file di folder source
@app.route('/source_file_count', methods=['GET'])
def source_file_count():
    source_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'source')
    skip_js = request.args.get('skip_js', 'false').lower() == 'true'
    
    if not os.path.exists(source_dir):
        return jsonify({"status": "error", "message": "Folder source tidak ditemukan"})
    
    files = []
    for root, _, filenames in os.walk(source_dir):
        for filename in filenames:
            if filename.endswith('.DAT'):
                if skip_js and 'JS' in filename:
                    continue
                files.append(os.path.join(root, filename))
    
    return jsonify({"status": "success", "count": len(files)})

# Route untuk mendapatkan progress pencarian
@app.route('/search_progress', methods=['GET'])
def search_progress():
    global CANCEL_SEARCH, SEARCH_RESULTS, SEARCH_TOTAL_FILES, SEARCH_PROCESSED_FILES
    
    with SEARCH_RESULTS_LOCK:
        count = len(SEARCH_RESULTS)
    
    return jsonify({
        "status": "success",
        "processed": SEARCH_PROCESSED_FILES,
        "total": SEARCH_TOTAL_FILES,
        "results": count,
        "is_cancelled": CANCEL_SEARCH
    })

# Route untuk membatalkan pencarian
@app.route('/cancel_search', methods=['POST'])
def cancel_search():
    global CANCEL_SEARCH, SEARCH_RESULTS, SEARCH_ACTIVE, SEARCH_COMPLETE
    
    CANCEL_SEARCH = True
    SEARCH_ACTIVE = False
    SEARCH_COMPLETE = True
    
    # Reset hasil pencarian juga
    with SEARCH_RESULTS_LOCK:
        SEARCH_RESULTS = []
    
    return jsonify({"status": "success", "message": "Pencarian dibatalkan dan hasil reset"})

# Route untuk reset hasil pencarian
@app.route('/reset_search', methods=['POST'])
def reset_search():
    global CANCEL_SEARCH, SEARCH_RESULTS, SEARCH_ACTIVE, SEARCH_COMPLETE
    
    CANCEL_SEARCH = True
    SEARCH_ACTIVE = False
    SEARCH_COMPLETE = True
    
    with SEARCH_RESULTS_LOCK:
        SEARCH_RESULTS = []
    
    return jsonify({"status": "success", "message": "Pencarian direset"})

# Route untuk mendapatkan hasil pencarian real-time
@app.route('/search_results', methods=['GET'])
def get_search_results():
    global SEARCH_RESULTS
    
    with SEARCH_RESULTS_LOCK:
        results = SEARCH_RESULTS.copy()
        total_results = len(SEARCH_RESULTS)
    
    # Cek apakah pencarian mungkin sudah selesai
    is_complete = total_results > 0
    
    print(f"get_search_results: Total={total_results}")
    
    return jsonify({
        "status": "success", 
        "count": total_results,
        "is_complete": is_complete,
        "results": results
    })

# Route untuk mendapatkan jumlah hasil tersimpan untuk Excel
@app.route('/excel_results_count', methods=['GET'])
def excel_results_count():
    global SEARCH_RESULTS
    
    with SEARCH_RESULTS_LOCK:
        count = len(SEARCH_RESULTS)
    
    return jsonify({
        "status": "success",
        "count": count
    })

# Route untuk upload folder source
@app.route('/upload_source', methods=['POST'])
def upload_source():
    if 'source_folder' not in request.files:
        return jsonify({"status": "error", "message": "Tidak ada file yang dipilih"})
    
    files = request.files.getlist('source_folder')
    
    if not files or files[0].filename == '':
        return jsonify({"status": "error", "message": "Tidak ada file yang dipilih"})
    
    # Hapus semua file yang ada di folder source
    for filename in os.listdir(app.config['SOURCE_FOLDER']):
        file_path = os.path.join(app.config['SOURCE_FOLDER'], filename)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
        except Exception as e:
            print(f"Error menghapus file {file_path}: {e}")
    
    file_count = 0
    for file in files:
        if file.filename:
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['SOURCE_FOLDER'], filename))
            file_count += 1
    
    return jsonify({
        "status": "success", 
        "message": f"{file_count} file berhasil diupload ke folder source"
    })

# Route untuk proses pencarian
@app.route('/search', methods=['POST'])
def search():
    global SEARCH_RESULTS, SEARCH_ACTIVE, SEARCH_COMPLETE, SEARCH_TOTAL_FILES, SEARCH_PROCESSED_FILES
    
    try:
        # Reset variabel status pencarian
        with SEARCH_RESULTS_LOCK:
            SEARCH_RESULTS = []
            SEARCH_ACTIVE = True
            SEARCH_COMPLETE = False
            SEARCH_TOTAL_FILES = 0
            SEARCH_PROCESSED_FILES = 0
        
        # Ambil input dari form
        table_names = request.form.getlist('table_name')
        remote_servers = request.form.getlist('remote_server')
        keywords = request.form.getlist('keywords')
        skip_js = request.form.get('skip_js', 'off') == 'on'
        
        # Ambil parameter overwrite_source
        overwrite_source = request.form.get('overwrite_source', 'false').lower() == 'true'
        
        # Simpan ke konfigurasi aplikasi
        app.config['OVERWRITE_SOURCE'] = overwrite_source
        
        # Tentukan number of workers
        max_workers = int(request.form.get('max_workers', '4'))
        
        # Jika "all" dipilih, ambil semua tabel
        if 'all' in table_names:
            table_names = [table_info['table_name'] for table_info in get_table_names()]
            remote_servers = [table_info['remote_server'] for table_info in get_table_names()]
            
            # Log untuk debugging
            print(f"Selected all tables: {len(table_names)} tables")
        
        # Validasi
        if not table_names or len(table_names) == 0:
            return jsonify({"status": "error", "message": "Pilih minimal satu nama tabel"})
        
        # Ambil source folder dari config
        source_folder = app.config.get('SOURCE_FOLDER')
        if not source_folder or not os.path.exists(source_folder):
            return jsonify({"status": "error", "message": "Source folder tidak valid"})
        
        # Ambil daftar file DAT
        files = []
        for file_name in os.listdir(source_folder):
            # Skip file JS jika dipilih
            if skip_js and file_name.lower().endswith('.js'):
                continue
            # Skip folder atau file non-reguler
            if os.path.isdir(os.path.join(source_folder, file_name)):
                continue
            files.append(file_name)
        
        # Log hasil yang didapat
        print(f"Processing {len(files)} files, {len(table_names)} tables, keywords: {keywords}, overwrite_source: {overwrite_source}")
        
        # Set total files untuk tracking progress
        with SEARCH_RESULTS_LOCK:
            SEARCH_TOTAL_FILES = len(files)
        
        # Override keywords jika array kosong, gunakan '-' sebagai wildcard (match any)
        if not keywords or len(keywords) == 0:
            keywords = ['-']
            
        # Jika lebih dari 1 keyword, lakukan pencarian terpisah untuk setiap keyword
        all_tasks = []
        for keyword in keywords:
            for file_name in files:
                all_tasks.append((file_name, table_names, remote_servers, keyword))

        # Load daftar semua tabel BE untuk pengecekan difficulty
        all_be_tables = load_tbl_be_tables()
        
        # Buat dictionary global untuk digunakan dalam fungsi process_file
        global all_tables, table_mapping
        all_tables = [table_info['table_name'] for table_info in get_table_names()]
        table_mapping = {table_info['table_name']: table_info.get('new_table_name', table_info['table_name']) for table_info in get_table_names()}
                
        # Update remote_servers berdasarkan tables yang dipilih
        remote_servers = [table_to_server.get(table, "") for table in table_names]
        
        # Buat thread utama untuk menjalankan semua pencarian dan menunggu sampai selesai
        search_thread = threading.Thread(
            target=process_search_tasks, 
            args=(all_tasks, max_workers)
        )
        search_thread.daemon = True
        search_thread.start()
        
        return jsonify({"status": "success", "message": "Pencarian dimulai", "task_count": len(all_tasks)})
    except Exception as e:
        # Tandai pencarian selesai jika terjadi exception
        with SEARCH_RESULTS_LOCK:
            SEARCH_ACTIVE = False
            SEARCH_COMPLETE = True
        
        error_message = f"Error: {str(e)}"
        print(error_message)
        return jsonify({"status": "error", "message": error_message})

# Route untuk cek status pencarian
@app.route('/search_status', methods=['GET'])
def get_search_status():
    global SEARCH_ACTIVE, SEARCH_COMPLETE, SEARCH_RESULTS
    
    with SEARCH_RESULTS_LOCK:
        count = len(SEARCH_RESULTS)
    
    return jsonify({
        "status": "success",
        "active": SEARCH_ACTIVE,
        "complete": SEARCH_COMPLETE,
        "count": count
    })

# Route untuk mengunduh hasil pencarian dalam format Excel
@app.route('/download_excel', methods=['GET'])
def download_excel():
    try:
        if not SEARCH_RESULTS or len(SEARCH_RESULTS) == 0:
            return jsonify({"status": "error", "message": "Tidak ada hasil pencarian"})
        
        # Buat timestamp untuk nama file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Buat workbook dan worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Search Results"
        
        # Tambahkan headers (disesuaikan sama seperti filtered)
        headers = ["No", "Nama File", "Nama Script", "Nomor Baris", "Tabel", "Remote Server", "Keyword", 
                  "Sudah Direplace", "Jenis Query", "Difficulty", "Script Sebelum", "Script Sesudah"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Tambahkan data
        for i, result in enumerate(SEARCH_RESULTS, 1):
            row_num = i + 1
            
            # Format query types menjadi string yang dapat dibaca
            query_types = result.get('query_type', [])
            if isinstance(query_types, list):
                query_type_str = ", ".join(query_types)
            else:
                query_type_str = str(query_types)
            
            # Tambahkan data ke excel (sama seperti filtered)
            ws.cell(row=row_num, column=1, value=i)  # No
            ws.cell(row=row_num, column=2, value=result.get('file_name', ''))  # Nama File
            ws.cell(row=row_num, column=3, value=result.get('script_name', ''))  # Nama Script
            ws.cell(row=row_num, column=4, value=result.get('line_number', ''))  # Nomor Baris
            ws.cell(row=row_num, column=5, value=result.get('table_name', ''))  # Tabel
            ws.cell(row=row_num, column=6, value=result.get('remote_server', ''))  # Remote Server
            ws.cell(row=row_num, column=7, value=result.get('keyword', ''))  # Keyword
            ws.cell(row=row_num, column=8, value='Ya' if result.get('replaced', False) else 'Tidak')  # Sudah Direplace
            ws.cell(row=row_num, column=9, value=query_type_str)  # Jenis Query
            ws.cell(row=row_num, column=10, value=result.get('difficulty', ''))  # Difficulty
            ws.cell(row=row_num, column=11, value=result.get('query', ''))  # Script Sebelum
            
            # Script Sesudah
            if result.get('replaced', False) and result.get('replaced_query', ''):
                ws.cell(row=row_num, column=12, value=result.get('replaced_query', ''))
            else:
                ws.cell(row=row_num, column=12, value='')
        
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2) if max_length < 80 else 80
            ws.column_dimensions[column].width = adjusted_width
        
        # Create a temporary file to save the Excel data
        temp_file = f"Tablix - {timestamp}.xlsx"
        wb.save(temp_file)
        
        return send_file(
            temp_file,
            as_attachment=True,
            download_name=f"Tablix - {timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return jsonify({"error": str(e)}), 500

# Route untuk mendapatkan semua hasil pencarian tanpa batasan
@app.route('/all_search_results', methods=['GET'])
def get_all_search_results():
    global SEARCH_RESULTS, SEARCH_ACTIVE, SEARCH_COMPLETE
    
    # Parameter opsional untuk menunggu sampai pencarian selesai
    wait_for_complete = request.args.get('wait_for_complete', 'false').lower() == 'true'
    
    # Jika diminta menunggu dan pencarian masih aktif, coba tunggu sebentar
    if wait_for_complete and SEARCH_ACTIVE and not SEARCH_COMPLETE:
        # Maksimal menunggu 3600 detik (sebelumnya 30 detik)
        wait_time = 0
        max_wait = 3600
        
        while SEARCH_ACTIVE and not SEARCH_COMPLETE and wait_time < max_wait:
            # Sleep 1 detik
            time.sleep(1)
            wait_time += 1
            
            # Log status
            with SEARCH_RESULTS_LOCK:
                print(f"Waiting for search to complete: {wait_time}/{max_wait} sec, {len(SEARCH_RESULTS)} results so far")
    
    # Selalu setel status pencarian ke selesai saat mengambil hasil
    SEARCH_ACTIVE = False
    SEARCH_COMPLETE = True
    
    # Menggunakan lock untuk memastikan konsistensi data
    with SEARCH_RESULTS_LOCK:
        results = SEARCH_RESULTS.copy()
        original_count = len(results)
    
    # Dapatkan mapping tabel ke tabel baru dari setting-tbl-name.txt
    table_mapping = {}
    try:
        with open('setting-tbl-name.txt', 'r') as file:
            for line in file:
                line = line.strip()
                if not line:
                    continue
                
                parts = line.split('\t')  # Coba dengan tab dulu
                
                if len(parts) == 1:
                    # Jika tidak ada tab, coba split dengan spasi
                    parts = line.split()
                
                if len(parts) >= 3:  # Pastikan ada setidaknya 3 kolom
                    table_name = parts[0].strip()
                    new_table_name = parts[2].strip()
                    table_mapping[table_name] = new_table_name
    except Exception as e:
        print(f"Error membaca setting-tbl-name.txt untuk mapping: {e}")
    
    # Dapatkan mapping nama script dari setting-script-name.txt
    script_names = get_script_names()
    
    # Tambahkan informasi new_table_name dan script_name ke hasil
    for result in results:
        # Pastikan kunci 'replaced' ada dalam setiap hasil
        if 'replaced' not in result:
            result['replaced'] = False
        
        table_name = result.get('table_name', '')
        result['new_table_name'] = table_mapping.get(table_name, table_name)
        
        # Tambahkan informasi nama script jika tersedia
        file_name = result.get('file_name', '')
        result['script_name'] = script_names.get(file_name, '')
        
        # Jika ada tables (multiple tables), tambahkan mapping untuk setiap tabel
        if 'tables' in result and len(result['tables']) > 0:
            result['new_table_names'] = []
            for tbl in result['tables']:
                result['new_table_names'].append(table_mapping.get(tbl, tbl))
    
    # Gabungkan hasil dengan query yang sama
    # Gunakan kombinasi query dan line_number sebagai kunci pengelompokan
    merged_results = []
    query_map = {}  # Map untuk melacak query yang sama di baris yang sama pada file yang sama
    
    for result in results:
        # Buat kunci gabungan dari file_name, query, dan line_number
        # Ini akan mengidentifikasi query SQL yang sama di baris yang sama pada file yang sama
        key = f"{result['file_name']}:{result['line_number']}:{hash(result['query'])}"
        
        if key in query_map:
            # Jika query ini sudah ada, tambahkan nama tabel ke daftar tabel yang ditemukan
            existing_result = query_map[key]
            
            # Gabungkan daftar tabel yang ditemukan
            if 'tables' not in existing_result:
                existing_result['tables'] = [existing_result['table_name']]
            
            # Tambahkan tabel baru jika belum ada dalam daftar
            if result['table_name'] not in existing_result['tables']:
                existing_result['tables'].append(result['table_name'])
            
            # Perbarui nama tabel menjadi daftar tabel yang ditemukan
            tabel_gabungan = ", ".join(existing_result['tables'])
            existing_result['table_name'] = tabel_gabungan
            
            # Gabungkan daftar new_table_names jika ada
            if 'new_table_names' not in existing_result and 'new_table_name' in existing_result:
                existing_result['new_table_names'] = [existing_result['new_table_name']]
            
            if 'new_table_name' in result:
                if result['new_table_name'] not in existing_result.get('new_table_names', []):
                    existing_result.setdefault('new_table_names', []).append(result['new_table_name'])
            
            # Update remote_server jika belum ada
            if not existing_result.get('remote_server', '') and result.get('remote_server', ''):
                existing_result['remote_server'] = result['remote_server']
            
            # Update needs_replace jika salah satu memerlukan replace
            if result.get('needs_replace', False) and not existing_result.get('needs_replace', False):
                existing_result['needs_replace'] = True
            
            # Update replaced jika salah satu sudah direplace
            if result.get('replaced', False) and not existing_result.get('replaced', False):
                existing_result['replaced'] = True
                
        else:
            # Jika ini query baru, tambahkan ke map
            # Pastikan kunci utama ada dalam result
            if 'tables' not in result:
                result['tables'] = [result['table_name']]
            
            # Pastikan new_table_names ada
            if 'new_table_name' in result and 'new_table_names' not in result:
                result['new_table_names'] = [result['new_table_name']]
                
            query_map[key] = result
            merged_results.append(result)
            
    # Kelompokkan hasil berdasarkan file_name
    file_groups = {}
    for result in merged_results:
        file_name = result.get('file_name', '')
        if file_name not in file_groups:
            file_groups[file_name] = []
        file_groups[file_name].append(result)
    
    # --- Tambahkan highlight per file ---
    source_folder = app.config.get('SOURCE_FOLDER')
    file_highlight_map = {}
    for file_name in file_groups.keys():
        file_path = os.path.join(source_folder, file_name)
        file_highlight_map[file_name] = scan_highlight_per_file(file_path)

    # Urutkan hasil dalam setiap grup berdasarkan nomor baris
    sorted_results = []
    for file_name, group in file_groups.items():
        # Urutkan berdasarkan nomor baris
        group.sort(key=lambda x: x.get('line_number', 0))
        # Tambahkan highlight di setiap result pertama per file
        if group:
            group[0]['file_highlights'] = file_highlight_map.get(file_name, [])
        sorted_results.extend(group)
    
    total_results = len(sorted_results)
    print(f"get_all_search_results: Returning {total_results} merged results (from {original_count} original)")
    
    return jsonify({
        "status": "success", 
        "count": original_count,  # Gunakan jumlah original sebelum penggabungan
        "search_active": False,  # Selalu kirim False untuk menandakan pencarian sudah selesai
        "search_complete": True, # Selalu kirim True untuk menandakan pencarian sudah selesai
        "results": sorted_results
    })

# Route untuk mengatur folder source dari path
@app.route('/set_source_path', methods=['POST'])
def set_source_path():
    source_path = request.form.get('source_path')
    
    if not source_path:
        return jsonify({"status": "error", "message": "Path folder tidak boleh kosong"})
    
    # Validasi path
    if not os.path.exists(source_path):
        return jsonify({"status": "error", "message": f"Path '{source_path}' tidak ditemukan"})
    
    if not os.path.isdir(source_path):
        return jsonify({"status": "error", "message": f"Path '{source_path}' bukan folder"})
    
    # Hitung jumlah file DAT dalam folder tersebut
    file_count = 0
    for root, _, filenames in os.walk(source_path):
        for filename in filenames:
            if filename.endswith('.DAT'):
                file_count += 1
    
    if file_count == 0:
        return jsonify({"status": "error", "message": "Tidak ada file DAT di folder tersebut"})
    
    # Set folder source ke path yang diberikan
    app.config['SOURCE_FOLDER'] = source_path
    
    print(f"Source folder diubah ke: {source_path} dengan {file_count} file DAT")
    
    return jsonify({
        "status": "success",
        "message": f"Folder source diatur ke '{source_path}' dengan {file_count} file DAT",
        "file_count": file_count
    })

# Route untuk mendapatkan isi file replacer.txt
@app.route('/get_replacer', methods=['GET'])
def get_replacer():
    try:
        with open('setting-remote-server.txt', 'r') as file:
            lines = file.readlines()
            replacer_data = []
            for line in lines:
                line = line.strip()
                if line:
                    parts = line.split(' ', 1)
                    remote_server = parts[0].strip()
                    description = parts[1].strip() if len(parts) > 1 else "-"
                    replacer_data.append({
                        "remote_server": remote_server,
                        "description": description
                    })
            return jsonify({"success": True, "data": replacer_data})
    except Exception as e:
        app.logger.error(f"Error reading setting-remote-server file: {str(e)}")
        return jsonify({"success": False, "error": "Failed to load remote server information"}), 500

# Route untuk mendapatkan remote servers
@app.route('/get_remote_servers', methods=['GET'])
def get_remote_servers():
    try:
        server_map = []
        with open('setting-remote-server.txt', 'r') as file:
            for line in file:
                line = line.strip()
                if line:
                    parts = line.split(' ', 1)
                    if len(parts) > 1:
                        server_map.append({
                            "remote_server": parts[0],
                            "description": parts[1]
                        })
                    else:
                        server_map.append({
                            "remote_server": parts[0],
                            "description": "-"
                        })
        return jsonify({"success": True, "data": server_map})
    except Exception as e:
        app.logger.error(f"Error reading setting-remote-server.txt: {str(e)}")
        return jsonify({"success": False, "error": "Failed to load remote server information"}), 500

# Fungsi untuk mencari file berdasarkan nama di dalam folder
def find_file_by_name(source_folder, file_name):
    """
    Mencari file dengan nama tertentu di dalam folder source.
    
    Args:
        source_folder: Path ke folder source
        file_name: Nama file yang dicari
        
    Returns:
        Path lengkap ke file jika ditemukan, None jika tidak ditemukan
    """
    for root, _, files in os.walk(source_folder):
        if file_name in files:
            return os.path.join(root, file_name)
    return None

# Route untuk mengganti teks 'jt' atau 'jdbctemplate' dalam file
@app.route('/replace_file', methods=['POST'])
def replace_file():
    try:
        data = request.get_json()
        
        # Validasi data yang diperlukan
        if 'file_name' not in data or 'line_number' not in data or 'remote_server' not in data:
            return jsonify({"success": False, "message": "Data tidak lengkap"}), 400
        
        file_name = data.get('file_name')
        line_number = int(data.get('line_number'))
        remote_server = data.get('remote_server')
        keyword = data.get('keyword', '').lower()
        original_query = data.get('query', '')  # Simpan query asli untuk dibandingkan
        overwrite_source = data.get('overwrite_source', False)  # Tambahan parameter baru
        
        # Ambil nama tabel jika ada
        table_names = data.get('table_names', [])
        new_table_names = data.get('new_table_names', [])
        
        # Cari file di folder source
        source_path = app.config.get('SOURCE_FOLDER', 'source')
        full_file_path = find_file_by_name(source_path, file_name)
        
        if not full_file_path:
            return jsonify({"success": False, "message": f"File {file_name} tidak ditemukan"}), 404
        
        # Periksa apakah folder source_replaced ada, jika tidak, buat folder tersebut
        source_dir = os.path.dirname(os.path.dirname(full_file_path))
        source_subpath = os.path.dirname(full_file_path).replace(source_path, '', 1).lstrip(os.sep)
        replaced_folder = os.path.join(source_dir, 'source_replaced')
        
        if not os.path.exists(replaced_folder):
            os.makedirs(replaced_folder)
            
        # Pastikan subfolder ada di source_replaced jika diperlukan
        if source_subpath:
            subfolder_path = os.path.join(replaced_folder, source_subpath)
            if not os.path.exists(subfolder_path):
                os.makedirs(subfolder_path)
        
        # Baca file
        with open(full_file_path, 'r', encoding='utf-8', errors='ignore') as file:
            lines = file.readlines()
        
        # Periksa apakah nomor baris valid
        if line_number < 1 or line_number > len(lines):
            return jsonify({"success": False, "message": f"Nomor baris {line_number} tidak valid"}), 400
        
        # Ambil konten line yang akan diganti
        original_line = lines[line_number - 1]
        modified_line = original_line
        
        # Baca setting-remote-server.txt untuk mendapatkan nilai yang benar untuk replacement
        replacement = None
        try:
            remote_server_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'setting-remote-server.txt')
            if os.path.exists(remote_server_file):
                with open(remote_server_file, 'r', encoding='utf-8', errors='ignore') as file:
                    for line in file:
                        parts = line.strip().split(' ', 1)  # Split by space, max 1 split
                        if len(parts) >= 2 and parts[0].strip() == remote_server:
                            replacement = parts[1].strip()
                            break
        except Exception as e:
            print(f"Error reading setting-remote-server.txt: {e}")
        
        # Jika replacement adalah '-' atau tidak ditemukan, jangan lakukan penggantian
        if replacement is None:
            return jsonify({"success": False, "message": f"Tidak ada nilai pengganti untuk server {remote_server}"}), 400
        
        if replacement == '-':
            return jsonify({"success": False, "message": f"Server {remote_server} ditandai dengan '-', tidak akan diganti"}), 400
        
        # Lakukan penggantian sesuai keyword
        # Tentukan metode getter untuk remoter server
        remote_server_suffix = ""
        if 'OLIBSCIF_DS' in remote_server:
            remote_server_suffix = "Cif"
        elif 'OLIBSFE_DS' in remote_server:
            remote_server_suffix = "Fe"
        elif 'OLIBSEXT_DS' in remote_server:
            remote_server_suffix = "Ext"
        elif 'OLIBSHTX_DS' in remote_server:
            remote_server_suffix = "Htx"
        elif 'OLIBSSBX_DS' in remote_server:
            remote_server_suffix = "Sbx"
        elif 'OLIBSSEC_DS' in remote_server:
            remote_server_suffix = "Sec"
        
        # Buat remote replacement berdasarkan suffix
        remote_replacement = f"dataSession.getJdbcTemplate{remote_server_suffix}()" if remote_server_suffix else replacement
        
        # Pastikan tidak ada karakter kontrol dalam replacement
        remote_replacement = ''.join(c for c in remote_replacement if ord(c) >= 32 or c == '\n' or c == '\t')
        
        # 1. Ganti pola jt. dengan dataSession.getJdbcTemplateXxx().
        if keyword == 'jt' or 'jt.' in keyword.lower():
            # Gunakan word boundary untuk memastikan bahwa kita hanya mengganti 'jt.' yang berdiri sendiri
            modified_line = re.sub(r'(\bjt\.|Jt\.|jT\.|JT\.)', f'{remote_replacement}.', modified_line)
            
        # Fungsi untuk memastikan tidak ada karakter kontrol dalam string
        def clean_string(s):
            return ''.join(c for c in s if ord(c) >= 32 or c == '\n' or c == '\t')
        
        # Pastikan remote_replacement bersih dari karakter kontrol
        clean_replacement = clean_string(remote_replacement)
        
        # Pola sederhana: hanya ganti 'jt' dengan batasan kata (word boundary)
        # Ini akan mengganti 'jt' yang berdiri sendiri sebagai kata, bukan bagian dari kata lain
        # Gunakan \b (word boundary) untuk memastikan hanya mengganti 'jt' yang berdiri sendiri
        jt_pattern = r'\bjt\b'
        modified_line = re.sub(jt_pattern, clean_replacement, modified_line)
        
        # Bersihkan hasil akhir dari karakter kontrol
        modified_line = clean_string(modified_line)
        
        # Ganti juga jdbcTemplate atau getJdbcTemplate() jika ada
        if 'jdbctemplate' in keyword.lower() or 'getjdbctemplate' in keyword.lower():
            # Tentukan metode getter untuk remoter server
            remote_server_suffix = ""
            if 'OLIBSCIF_DS' in remote_server:
                remote_server_suffix = "Cif"
            elif 'OLIBSFE_DS' in remote_server:
                remote_server_suffix = "Fe"
            elif 'OLIBSEXT_DS' in remote_server:
                remote_server_suffix = "Ext"
            elif 'OLIBSHTX_DS' in remote_server:
                remote_server_suffix = "Htx"
            elif 'OLIBSSBX_DS' in remote_server:
                remote_server_suffix = "Sbx"
            elif 'OLIBSSEC_DS' in remote_server:
                remote_server_suffix = "Sec"
            
            # Jika ada suffix, gunakan metode getter yang sesuai
            if remote_server_suffix:
                remote_replacement = f"dataSession.getJdbcTemplate{remote_server_suffix}()"
                # Perbaiki penggantian getJdbcTemplate() agar tidak menghasilkan kurung ganda
                # Untuk pola getJdbcTemplate().method() - ganti dengan dataSession.getJdbcTemplateFe().method()
                modified_line = re.sub(r'\bgetJdbcTemplate\(\)\.', f'{remote_replacement}.', modified_line, flags=re.IGNORECASE)
                
                # Untuk pola getJdbcTemplate() tanpa apa-apa setelahnya - ganti dengan dataSession.getJdbcTemplateFe()
                modified_line = re.sub(r'\bgetJdbcTemplate\(\)(?!\.)(?=\s|$|\)|\,)', f'{remote_replacement}', modified_line, flags=re.IGNORECASE)
                
                # Cari dan ganti secara lebih spesifik dengan word boundary untuk jdbcTemplate
                modified_line = re.sub(r'\bjdbcTemplate\b', f'{remote_replacement}', modified_line, flags=re.IGNORECASE)
            else:
                # Fallback menggunakan replacement dari setting-remote-server.txt
                modified_line = re.sub(r'\bgetJdbcTemplate\(\)\.', f'{replacement}.', modified_line, flags=re.IGNORECASE)
                modified_line = re.sub(r'\bgetJdbcTemplate\(\)(?!\.)(?=\s|$|\)|\,)', f'{replacement}', modified_line, flags=re.IGNORECASE)
                modified_line = re.sub(r'\bjdbcTemplate\b', replacement, modified_line, flags=re.IGNORECASE)
        
        # Cari pasangan tabel BE dan tabel Remote Server berdasarkan remote_server
        table_mapping = {}
        try:
            tbl_name_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'setting-tbl-name.txt')
            if os.path.exists(tbl_name_file):
                with open(tbl_name_file, 'r', encoding='utf-8', errors='ignore') as file:
                    for line in file:
                        parts = line.strip().split('\t')
                        if len(parts) >= 3 and parts[1].strip() == remote_server:
                            be_table_name = parts[0].strip()
                            remote_table_name = parts[2].strip()
                            table_mapping[be_table_name] = remote_table_name
        except Exception as e:
            print(f"Error reading setting-tbl-name.txt: {e}")
        
        # Ganti nama tabel BE dengan nama tabel di Remote Server
        if table_mapping:
            # Urutkan tabel berdasarkan panjang (terpanjang lebih dulu) untuk menghindari penggantian substring
            for be_table, remote_table in sorted(table_mapping.items(), key=lambda x: len(x[0]), reverse=True):
                if be_table != remote_table:  # Hanya ganti jika nama tabelnya berbeda
                    # Pastikan mencari tabel dengan batasan kata agar tidak mengganti substring
                    modified_line = re.sub(fr'\b{re.escape(be_table)}\b', remote_table, modified_line)
        
        # Ganti nama tabel jika diperlukan (tambahan dari parameter)
        if table_names and new_table_names and len(table_names) == len(new_table_names):
            for i, table_name in enumerate(table_names):
                # Pastikan mencari tabel dengan batasan kata agar tidak mengganti substring
                modified_line = re.sub(fr'\b{re.escape(table_name)}\b', new_table_names[i], modified_line)
        
        # Jika tidak ada perubahan, kembalikan error
        if modified_line == original_line:
            return jsonify({"success": False, "message": "Tidak ada perubahan yang dilakukan"}), 400
        
        # Simpan file yang telah diubah
        lines[line_number - 1] = modified_line
        
        # Simpan file berdasarkan flag overwrite_source
        if overwrite_source:
            # Jika overwrite_source, hanya simpan ke folder source asli
            with open(full_file_path, 'w', encoding='utf-8', errors='ignore') as file:
                file.writelines(lines)
            print(f"File {file_name} ditimpa di folder source asli: {full_file_path}")
        else:
            # Jika tidak overwrite_source, simpan ke folder source_replaced dengan menjaga struktur direktori
            if source_subpath:
                # Pastikan subfolder ada di source_replaced
                subfolder_path = os.path.join(replaced_folder, source_subpath)
                if not os.path.exists(subfolder_path):
                    os.makedirs(subfolder_path)
                replaced_file_path = os.path.join(replaced_folder, source_subpath, os.path.basename(full_file_path))
            else:
                replaced_file_path = os.path.join(replaced_folder, os.path.basename(full_file_path))
            
            # Periksa apakah file sudah ada di folder source_replaced
            # Jika sudah ada, baca file tersebut terlebih dahulu agar tidak menimpa perubahan sebelumnya
            if os.path.exists(replaced_file_path):
                try:
                    with open(replaced_file_path, 'r', encoding='utf-8', errors='ignore') as existing_file:
                        existing_lines = existing_file.readlines()
                    
                    # Jika panjang baris sama, ganti baris yang dimodifikasi di file yang ada
                    if len(existing_lines) == len(lines):
                        existing_lines[line_number - 1] = modified_line
                        lines = existing_lines
                    # Jika panjang baris berbeda, gunakan file baru dengan baris yang dimodifikasi
                except Exception as e:
                    print(f"Error reading existing replaced file: {e}")
                    # Lanjutkan dengan file baru jika terjadi error
            
            # Tulis ke file di folder source_replaced
            with open(replaced_file_path, 'w', encoding='utf-8', errors='ignore') as file:
                file.writelines(lines)
            print(f"File {file_name} disimpan di folder source_replaced: {replaced_file_path}")
        
        # Update status di SEARCH_RESULTS
        with SEARCH_RESULTS_LOCK:
            for result in SEARCH_RESULTS:
                if (result.get('file_name') == file_name and 
                    int(result.get('line_number')) == line_number):
                    result['replaced'] = True
                    # Simpan query yang sudah direplace untuk Excel
                    result['replaced_query'] = modified_line.strip()
        
        # Kembalikan respons sukses
        return jsonify({
            "success": True, 
            "message": "File berhasil direplace",
            "file": file_name,
            "line": line_number,
            "original_line": original_line,
            "new_line": modified_line,
            "replacement_value": replacement
        })
    except Exception as e:
        print(f"Error in replace_file: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# Error handlers
@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html', error="403 Forbidden: Akses ditolak"), 403

@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', error="404 Not Found: Halaman tidak ditemukan"), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('error.html', error="500 Internal Server Error: Terjadi kesalahan di server"), 500

@app.route('/test_connection')
def test_connection():
    """Endpoint untuk menguji koneksi."""
    return jsonify({"status": "success", "message": "Koneksi berhasil!"})

def calculate_difficulty(tables, query, all_be_tables):
    """
    Menentukan tingkat kesulitan query berdasarkan jumlah tabel dan keberadaan tabel BE.
    
    Args:
        tables: List tabel yang ditemukan
        query: Query SQL
        all_be_tables: Semua tabel BE dari setting-tbl-be.txt
    
    Returns:
        String: 'easy', 'medium', atau 'hard'
    """
    # Jika hanya ada 1 tabel, difficulty easy
    if len(tables) == 1:
        return 'easy'
    
    # Jika ada lebih dari 1 tabel, difficulty medium
    if len(tables) > 1:
        return 'medium'
    
    # Cek apakah ada tabel BE yang muncul di query
    query_lower = query.lower()
    be_tables_in_query = [table for table in all_be_tables if table.lower() in query_lower and table.lower() not in [t.lower() for t in tables]]
    
    # Jika ada tabel BE dalam query tapi tidak terdeteksi di 'tables', maka hard
    if be_tables_in_query:
        return 'hard'
    
    # Default ke medium jika tidak ada kondisi di atas yang terpenuhi
    return 'medium'

@app.route('/download_excel_filtered', methods=['POST'])
def download_excel_filtered():
    try:
        # Get filtered data from request
        data = request.get_json()
        if not data or 'results' not in data or not data['results']:
            return jsonify({"error": "No filtered results provided"}), 400
        
        filtered_results = data['results']
        
        # Buat timestamp untuk nama file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Buat workbook dan worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered Results"
        
        # Tambahkan headers (menghapus "Memerlukan Replace" dan mengubah "Query" menjadi "Script Sebelum")
        headers = ["No", "Nama File", "Nama Script", "Nomor Baris", "Tabel", "Remote Server", "Keyword", 
                  "Sudah Direplace", "Jenis Query", "Difficulty", "Script Sebelum", "Script Sesudah"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Tambahkan data
        for i, result in enumerate(filtered_results, 1):
            row_num = i + 1
            
            # Format query types menjadi string yang dapat dibaca
            query_types = result.get('query_type', [])
            if isinstance(query_types, list):
                query_type_str = ", ".join(query_types)
            else:
                query_type_str = str(query_types)
            
            # Tambahkan data ke excel
            ws.cell(row=row_num, column=1, value=i)  # No
            ws.cell(row=row_num, column=2, value=result.get('file_name', ''))  # Nama File
            ws.cell(row=row_num, column=3, value=result.get('script_name', ''))  # Nama Script
            ws.cell(row=row_num, column=4, value=result.get('line_number', ''))  # Nomor Baris
            ws.cell(row=row_num, column=5, value=result.get('table_name', ''))  # Tabel
            ws.cell(row=row_num, column=6, value=result.get('remote_server', ''))  # Remote Server
            ws.cell(row=row_num, column=7, value=result.get('keyword', ''))  # Keyword
            # Kolom "Sudah Direplace" (gunakan hasil dari UI, result.replaced)
            ws.cell(row=row_num, column=8, value='Ya' if result.get('replaced', False) else 'Tidak')
            ws.cell(row=row_num, column=9, value=query_type_str)  # Jenis Query
            ws.cell(row=row_num, column=10, value=result.get('difficulty', ''))  # Difficulty
            ws.cell(row=row_num, column=11, value=result.get('query', ''))  # Script Sebelum
            
            # Script Sesudah (kosong jika belum direplace, berisi query setelah direplace jika sudah)
            if result.get('replaced', False) and result.get('replaced_query', ''):
                ws.cell(row=row_num, column=12, value=result.get('replaced_query', ''))
            else:
                ws.cell(row=row_num, column=12, value='')
        
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2) if max_length < 80 else 80
            ws.column_dimensions[column].width = adjusted_width
        
        # Create a temporary file to save the Excel data
        temp_file = f"Tablix Filtered - {timestamp}.xlsx"
        wb.save(temp_file)
        
        return send_file(
            temp_file,
            as_attachment=True,
            download_name=f"Tablix Filtered - {timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"Error exporting filtered results to Excel: {e}")
        return jsonify({"error": str(e)}), 500

def check_already_replaced(line, remote_server=None):
    """
    Mengecek apakah baris sudah direplace berdasarkan nilai di setting-remote-server.txt
    
    Args:
        line: Baris yang akan dicek
        remote_server: Server yang terkait (opsional, jika None akan cek semua server)
        
    Returns:
        bool: True jika sudah direplace, False jika belum
    """
    try:
        # Baca semua replacement value dari setting-remote-server.txt
        replacements = []
        remote_server_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'setting-remote-server.txt')
        
        if os.path.exists(remote_server_file):
            with open(remote_server_file, 'r', encoding='utf-8', errors='ignore') as file:
                for rs_line in file:
                    parts = rs_line.strip().split(' ', 1)
                    if len(parts) >= 2 and parts[1] != '-':
                        rs_name = parts[0].strip()
                        rs_replacement = parts[1].strip()
                        
                        # Jika remote_server diberikan, hanya cek untuk server tersebut
                        if remote_server is None or rs_name == remote_server:
                            replacements.append(rs_replacement)
        
        # Logic yang diperbaiki: data dianggap sudah direplace jika mengandung
        # kata-kata dari setting-remote-server.txt di kolom 2 (kecuali yang isinya '-')
        for replacement in replacements:
            if replacement in line:
                return True
                
        return False
    except Exception as e:
        print(f"Error dalam check_already_replaced: {e}")
        return False

def process_file(file_info):
    """
    Fungsi helper untuk memproses masing-masing file secara paralel.
    """
    try:
        file_name, tables, remote_servers, keyword = file_info

        # Baca file
        file_path = os.path.join(app.config['SOURCE_FOLDER'], file_name)
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.readlines()
        
        # Cari nama script untuk file ini
        script_name = get_script_name(file_name)
        
        # Siapkan hasil untuk file ini
        file_results = []
        
        # Iterasi setiap baris dalam file
        for i, line in enumerate(content):
            for idx, table in enumerate(tables):
                remote_server = remote_servers[idx] if idx < len(remote_servers) else ""
                
                # Lewati jika table kosong
                if not table:
                    continue
                
                # Cek apakah table ada dalam baris dengan exact match word boundary
                table_pattern = r'(^|[\s\'"`\(\),;])' + re.escape(table) + r'($|[\s\'"`\(\),;])'
                if not re.search(table_pattern, line, re.IGNORECASE):
                    continue
                
                # Cek juga kita memfilter keyword (jika ada)
                keyword_found = True
                if keyword:
                    # Jika keyword adalah '-', anggap selalu ditemukan (wildcard)
                    if keyword == '-':
                        keyword_found = True
                    else:
                        keyword_found = keyword.lower() in line.lower()
                
                if not keyword_found:
                    continue
                
                # Cek apakah script sudah direplace berdasarkan isi setting-remote-server.txt
                already_replaced = check_already_replaced(line, remote_server)
                
                # Tentukan apakah keyword jt atau jdbctemplate perlu direplace
                needs_replace = False
                if keyword and remote_server and not already_replaced:
                    keyword_lower = keyword.lower()
                    if (keyword_lower == 'jt' and ('jt.' in line.lower() or 'jdbctemplate' in line.lower() or 'getjdbctemplate()' in line.lower())):
                        needs_replace = True
                    elif keyword_lower in ['jdbctemplate', 'getjdbctemplate', 'getjdbctemplate()'] and ('jdbctemplate' in line.lower() or 'getjdbctemplate' in line.lower()):
                        needs_replace = True
                
                # Dapatkan nama tabel baru jika tersedia
                new_table_name = table_mapping.get(table, table)
                
                # Cari tabel lain yang mungkin juga ada di query yang sama
                # Ini untuk kasus multi-table query
                other_tables_in_line = []
                new_table_names = []
                
                for t in all_tables:
                    if t != table:
                        # Gunakan exact match dengan boundary untuk tabel
                        other_table_pattern = r'(^|[\s\'"`\(\),;])' + re.escape(t) + r'($|[\s\'"`\(\),;])'
                        if re.search(other_table_pattern, line, re.IGNORECASE):
                            other_tables_in_line.append(t)
                            # Tambahkan nama tabel baru juga
                            new_table_names.append(table_mapping.get(t, t))
                
                tables_in_this_line = [table] + other_tables_in_line
                new_table_names = [new_table_name] + new_table_names
                
                # Tentukan jenis query: select atau manipulation
                query_types = []
                
                # Deteksi SELECT queries
                select_query_methods = ['query(', 'queryobject(', 'queryforobject(', 'createquery(', 'queryforint(', 'queryforstring(', 'queryforbigdecimal(']
                
                # Deteksi SELECT queries
                is_select_query = False
                
                # Cek apakah ada method SELECT
                if any(method.lower() in line.lower() for method in select_query_methods):
                    is_select_query = True
                # Cek apakah ada kata SELECT dalam konteks query SQL
                elif 'select' in line.lower() and ('from' in line.lower() or '=' in line.lower()):
                    is_select_query = True
                    
                if is_select_query:
                    query_types.append('select')
                    
                # Deteksi MANIPULATION queries
                manipulation_keywords = [
                    'insert', 'update', 'delete',
                    'insertdata', 'updatedata', 'deletedata',
                    '.addsql(', '.execute(',
                    'updatedataquery', 'insertdataquery', 'deletedataquery'
                ]
                
                is_manipulation_query = any(keyword in line.lower() for keyword in manipulation_keywords)
                
                # Jika tidak ada query_types yang terdeteksi, cek apakah ada pola umum untuk manipulasi data
                if not is_manipulation_query and ('jt,' in line.lower() or 'jt"' in line.lower() or "jt'" in line.lower()):
                    is_manipulation_query = True
                
                # Jika terdeteksi sebagai manipulasi, tambahkan ke query_types
                if is_manipulation_query:
                    query_types.append('manipulation')
                    
                # Reset query_types jika terdeteksi sebagai SELECT+MANIPULASI berdasarkan definisi yang tepat
                # SELECT+MANIPULASI: ada kata 'select' dan salah satu atau lebih dari kata 'update', 'insert', atau 'delete'
                line_lower = line.lower()
                has_select = 'select' in line_lower
                has_manipulation = 'update' in line_lower or 'insert' in line_lower or 'delete' in line_lower
                
                # Jika dalam satu baris ada SELECT dan salah satu kata manipulasi (update, insert, delete),
                # maka ini adalah SELECT+MANIPULASI, bukan SELECT dan MANIPULASI terpisah
                if has_select and has_manipulation and len(query_types) > 1:
                    query_types = ['select+manipulation']
                
                # Hitung difficulty berdasarkan kriteria
                difficulty = 'medium'  # Nilai default
                
                # 1. Jika hanya ada 1 tabel, maka Easy
                if len(tables_in_this_line) == 1:
                    difficulty = 'easy'
                # 2. Jika ada lebih dari 1 tabel, maka Medium
                elif len(tables_in_this_line) > 1:
                    difficulty = 'medium'
                
                # 3. Cek apakah ada tabel dari setting-tbl-be.txt yang muncul di query tapi tidak terdeteksi di tables_in_this_line
                query_lower = line.lower()
                detected_table_names_lower = [t.lower() for t in tables_in_this_line]
                
                # Cek apakah ada tabel BE yang muncul di query tapi tidak terdeteksi
                for be_table in all_be_tables:
                    # Gunakan exact match dengan word boundary
                    be_table_pattern = r'(^|[\s\'"`\(\),;])' + re.escape(be_table.lower()) + r'($|[\s\'"`\(\),;])'
                    if re.search(be_table_pattern, query_lower) and be_table.lower() not in detected_table_names_lower:
                        difficulty = 'hard'
                        break
                
                # Hasil pencarian ditemukan, tambahkan ke list
                file_results.append({
                    'file_name': file_name,
                    'script_name': script_name,
                    'line_number': i + 1,
                    'query': line.strip(),
                    'table_name': table,
                    'tables': tables_in_this_line,
                    'new_table_name': new_table_name,
                    'new_table_names': new_table_names,
                    'remote_server': remote_server,
                    'keyword': keyword,
                    'needs_replace': needs_replace,
                    'replaced': already_replaced,  # Set replaced = True jika sudah direplace
                    'query_type': query_types,
                    'difficulty': difficulty
                })
        
        return file_results
    except Exception as e:
        print(f"Error processing file {file_info[0]}: {e}")
        return []

def process_search_tasks(all_tasks, max_workers):
    """
    Fungsi untuk memproses semua tugas pencarian menggunakan worker pool.
    
    Args:
        all_tasks: List tugas pencarian [(file_name, tables, remote_servers, keyword), ...]
        max_workers: Jumlah maksimum worker
    """
    global SEARCH_RESULTS, SEARCH_ACTIVE, SEARCH_COMPLETE, SEARCH_PROCESSED_FILES
    
    try:
        # Buat executor
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all tasks and collect futures
            futures = [executor.submit(process_file, task) for task in all_tasks]
            
            # Process results as they complete
            for future in as_completed(futures):
                with SEARCH_RESULTS_LOCK:
                    SEARCH_PROCESSED_FILES += 1
                    
                    # Get result from this future
                    results = future.result()
                    if results:
                        SEARCH_RESULTS.extend(results)
                    
                    # Print progress
                    print(f"Processed {SEARCH_PROCESSED_FILES}/{len(all_tasks)} files, "
                          f"{len(SEARCH_RESULTS)} results so far")
        
        # Mark search as complete
        with SEARCH_RESULTS_LOCK:
            SEARCH_COMPLETE = True
            SEARCH_ACTIVE = False
            print(f"Search completed with {len(SEARCH_RESULTS)} total results")
            
    except Exception as e:
        print(f"Error in process_search_tasks: {e}")
        with SEARCH_RESULTS_LOCK:
            SEARCH_COMPLETE = True
            SEARCH_ACTIVE = False

@app.route('/set_intellij_path', methods=['POST'])
def set_intellij_path():
    intellij_path = request.form.get('intellij_path')
    if not intellij_path:
        return jsonify({"status": "error", "message": "Path Intellij tidak boleh kosong"})
    if not os.path.exists(intellij_path):
        return jsonify({"status": "error", "message": f"Path '{intellij_path}' tidak ditemukan"})
    if not os.path.isdir(intellij_path):
        return jsonify({"status": "error", "message": f"Path '{intellij_path}' bukan folder"})
    app.config['INTELLIJ_PATH'] = intellij_path
    print(f"Intellij path diubah ke: {intellij_path}")
    return jsonify({
        "status": "success",
        "message": f"Path Intellij diatur ke '{intellij_path}'"
    })

@app.route('/goto_code', methods=['POST'])
def goto_code():
    data = request.get_json()
    file_name = data.get('file_name')
    line_number = data.get('line_number')
    if not file_name or not line_number:
        return jsonify({"success": False, "message": "Data tidak lengkap"}), 400
    intellij_path = app.config.get('INTELLIJ_PATH')
    if not intellij_path:
        return jsonify({"success": False, "message": "Path Intellij belum diatur"}), 400
    file_path = os.path.join(intellij_path, file_name)
    if not os.path.exists(file_path):
        return jsonify({"success": False, "message": f"File {file_path} tidak ditemukan"}), 404
    import subprocess
    try:
        subprocess.Popen(["idea", f"--line", str(line_number), file_path])
        return jsonify({"success": True, "message": f"Berhasil membuka {file_name} di Intellij pada baris {line_number}"})
    except Exception as e:
        return jsonify({"success": False, "message": f"Gagal menjalankan perintah: {e}"}), 500

def scan_highlight_per_file(file_path):
    """
    Scan file untuk mencari baris yang mengandung generateForm dan execute/executes/jt.executes.
    Return list of dict: {type, value, line}
    """
    highlights = []
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        for idx, line in enumerate(lines):
            # Cari generateForm
            match = re.search(r'generateForm\(([^)]*)\)', line)
            if match:
                param = match.group(1).strip().strip('"\'')
                highlights.append({
                    'type': 'generateForm',
                    'value': param,
                    'line': idx + 1
                })
            # Cari jt.executes, executes, execute (tapi bukan bagian dari kata lain)
            for exec_type in ['jt.executes', 'executes', 'execute']:
                # Pastikan bukan bagian dari kata lain (pakai boundary)
                if re.search(r'\b' + re.escape(exec_type) + r'\b', line):
                    highlights.append({
                        'type': 'execute',
                        'value': exec_type,
                        'line': idx + 1
                    })
    except Exception as e:
        print(f"Error scan_highlight_per_file {file_path}: {e}")
    return highlights

if __name__ == '__main__':
    # Jalankan pada semua interface (0.0.0.0) dan port 5000
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True) 